"""
Index manufacturing domain knowledge from Shipment_Stock_Sheet.xlsx into
Cairn's pgvector store for hybrid BM25 + cosine similarity retrieval.

Chunks indexed:
  - product: one per M-number (description, blank, material, stock)
  - sku_mapping: SKU→M-number mappings per channel
  - stock_level: optimal stock levels from ScratchPad2
  - material: procurement materials and suppliers

Usage:
  python scripts/index_manufacturing_domain.py [--file path/to/xlsx]
"""
import argparse
import hashlib
import re
import sys
import time

import psycopg2
import requests
from openpyxl import load_workbook

OLLAMA_BASE = 'http://localhost:11434'
EMBED_MODEL = 'nomic-embed-text'
PROJECT_ID = 'manufacturing'
DB_CONFIG = {
    'dbname': 'claw',
    'user': 'postgres',
    'password': 'postgres123',
    'host': 'localhost',
    'port': '5432',
}

UPSERT_SQL = """
DELETE FROM claw_code_chunks WHERE project_id = %s AND content_hash = %s;
INSERT INTO claw_code_chunks
    (project_id, file_path, chunk_content, chunk_type, chunk_name,
     content_hash, embedding, subproject_id, indexed_at)
VALUES (%s, %s, %s, %s, %s, %s, %s::vector, %s, NOW())
"""


def embed(text: str) -> list[float] | None:
    text = text[:1500]
    try:
        resp = requests.post(
            f'{OLLAMA_BASE}/api/embeddings',
            json={'model': EMBED_MODEL, 'prompt': text},
            timeout=30,
        )
        resp.raise_for_status()
        return resp.json()['embedding']
    except Exception as e:
        print(f'  [WARN] Embedding failed: {e}')
        return None


def content_hash(text: str) -> str:
    return hashlib.sha256(text.encode()).hexdigest()


def index_master_stock(wb, conn):
    ws = wb['MASTER STOCK']
    rows = list(ws.iter_rows(values_only=True))
    header = rows[2]

    col_map = {}
    for idx, val in enumerate(header):
        if val:
            col_map[str(val).strip().upper()] = idx

    count = 0
    cur = conn.cursor()

    for row in rows[3:]:
        m = row[col_map.get('MASTER', 0)]
        if not m or not re.match(r'^M\d+$', str(m).strip().upper()):
            continue

        m = str(m).strip().upper()
        desc = str(row[col_map.get('DESCRIPTION', 1)] or '').strip()
        blank = str(row[col_map.get('BLANK', 2)] or '').strip()
        material = str(row[col_map.get('MATERIAL', 3)] or '').strip()
        stock = row[col_map.get('STOCK', 4)] or 0

        if not desc:
            continue

        text = f'Product {m}: {desc}. Blank: {blank}. Material: {material}. Stock: {stock}.'
        h = content_hash(text)
        vec = embed(text)
        if not vec:
            continue

        cur.execute(UPSERT_SQL, (
            PROJECT_ID, h,
            PROJECT_ID, f'spreadsheet/MASTER_STOCK/{m}', text,
            'product', m, h, str(vec), None,
        ))
        count += 1

        if count % 100 == 0:
            conn.commit()
            print(f'  Products indexed: {count}')

    conn.commit()
    print(f'  Total products indexed: {count}')
    return count


def index_assembly(wb, conn):
    ws = wb['ASSEMBLY']
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    col_map = {}
    for idx, val in enumerate(header):
        if val:
            col_map[str(val).strip().upper()] = idx

    count = 0
    cur = conn.cursor()

    for row in rows[1:]:
        sku = row[col_map.get('SKU', 0)]
        master = row[col_map.get('MASTER SKU', 1)]
        if not sku or not master:
            continue

        sku = str(sku).strip()
        master = str(master).strip().upper()
        if not re.match(r'^M\d+$', master):
            continue

        country = str(row[col_map.get('COUNTRY', 3)] or '').strip()
        desc = str(row[col_map.get('DESCRIPTION', 4)] or '').strip()
        blank = str(row[col_map.get('BLANK', 5)] or '').strip()
        asin = str(row[col_map.get('ASIN', 7)] or '').strip()

        text = f'SKU {sku} maps to {master} ({country}). {desc}. Blank: {blank}. ASIN: {asin}.'
        h = content_hash(text)
        vec = embed(text)
        if not vec:
            continue

        cur.execute(UPSERT_SQL, (
            PROJECT_ID, h,
            PROJECT_ID, f'spreadsheet/ASSEMBLY/{sku}', text,
            'sku_mapping', f'{master}/{country}/{sku}', h, str(vec), None,
        ))
        count += 1

        if count % 200 == 0:
            conn.commit()
            print(f'  SKU mappings indexed: {count}')

    conn.commit()
    print(f'  Total SKU mappings indexed: {count}')
    return count


def index_scratchpad(wb, conn):
    ws = wb['ScratchPad2']
    rows = list(ws.iter_rows(values_only=True))

    count = 0
    cur = conn.cursor()

    for row in rows[1:]:
        m = row[0]
        optimal = row[1]
        if not m:
            continue

        m = str(m).strip().upper()
        if not re.match(r'^M\d+$', m):
            continue

        optimal = int(optimal) if optimal else 0
        text = f'Optimal 30-day stock for {m}: {optimal} units.'
        h = content_hash(text)
        vec = embed(text)
        if not vec:
            continue

        cur.execute(UPSERT_SQL, (
            PROJECT_ID, h,
            PROJECT_ID, f'spreadsheet/ScratchPad2/{m}', text,
            'stock_level', m, h, str(vec), None,
        ))
        count += 1

    conn.commit()
    print(f'  Total stock levels indexed: {count}')
    return count


def index_procurement(wb, conn):
    ws = wb['PROCUREMENT']
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    col_map = {}
    for idx, val in enumerate(header):
        if val:
            col_map[str(val).strip()] = idx

    count = 0
    cur = conn.cursor()

    for row in rows[1:]:
        mat_id = row[col_map.get('MaterialID', 0)]
        if not mat_id:
            continue

        name = str(row[col_map.get('MaterialName', 1)] or '').strip()
        category = str(row[col_map.get('Category', 2)] or '').strip()
        supplier = str(row[col_map.get('PreferredSupplierID', 7)] or '').strip()

        text = f'Material {mat_id}: {name}. Category: {category}. Supplier: {supplier}.'
        h = content_hash(text)
        vec = embed(text)
        if not vec:
            continue

        cur.execute(UPSERT_SQL, (
            PROJECT_ID, h,
            PROJECT_ID, f'spreadsheet/PROCUREMENT/{mat_id}', text,
            'material', str(mat_id), h, str(vec), None,
        ))
        count += 1

    conn.commit()
    print(f'  Total materials indexed: {count}')
    return count


def main():
    parser = argparse.ArgumentParser(description='Index manufacturing domain into Cairn')
    parser.add_argument('--file', default='C:/Users/zentu/Downloads/Shipment Stock Sheet (1).xlsx')
    args = parser.parse_args()

    print(f'Checking Ollama embedding model ({EMBED_MODEL})...')
    try:
        test = embed('test')
        if not test:
            print('ERROR: Embedding model not available. Run: ollama pull nomic-embed-text')
            sys.exit(1)
        print(f'  OK — {len(test)} dimensions')
    except Exception as e:
        print(f'ERROR: Cannot reach Ollama: {e}')
        sys.exit(1)

    print(f'Loading workbook: {args.file}')
    wb = load_workbook(args.file, read_only=True, data_only=True)

    print('Connecting to Cairn database...')
    conn = psycopg2.connect(**DB_CONFIG)

    start = time.time()
    totals = {}

    print('\n--- Indexing MASTER STOCK ---')
    totals['products'] = index_master_stock(wb, conn)

    print('\n--- Indexing ASSEMBLY ---')
    totals['sku_mappings'] = index_assembly(wb, conn)

    print('\n--- Indexing ScratchPad2 ---')
    totals['stock_levels'] = index_scratchpad(wb, conn)

    print('\n--- Indexing PROCUREMENT ---')
    totals['materials'] = index_procurement(wb, conn)

    elapsed = time.time() - start
    wb.close()
    conn.close()

    print(f'\n=== COMPLETE in {elapsed:.1f}s ===')
    for k, v in totals.items():
        print(f'  {k}: {v} chunks')
    print(f'  Total: {sum(totals.values())} chunks indexed into project "{PROJECT_ID}"')


if __name__ == '__main__':
    main()
