"""
Import historical FBA shipments from Previous FBA Shipments sheet.
Groups rows by shipment date + country into Shipment records.
Columns: Country, SKU, M SKU, NOTES, Name, Owner, AMZ Restock Quantity,
         MAKE?, M STOCK, Number Shipped, [shipment date text in col 10]
"""
import re
from collections import defaultdict
from datetime import datetime
from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl import load_workbook
from products.models import Product
from shipments.models import Shipment, ShipmentItem
from imports.models import ImportLog


def parse_ship_date(text):
    if not text:
        return None
    text = str(text).strip()
    m = re.search(r'(\d{2}/\d{2}/\d{2})', text)
    if m:
        try:
            return datetime.strptime(m.group(1), '%d/%m/%y').date()
        except ValueError:
            pass
    return None


class Command(BaseCommand):
    help = 'Import historical FBA shipments from Previous FBA Shipments sheet'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, default=None)
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        filepath = options['file'] or settings.SPREADSHEET_PATH
        dry_run = options['dry_run']

        self.stdout.write(f'Loading {filepath}...')
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb['Previous FBA Shipments']

        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]

        col_map = {}
        for idx, val in enumerate(header):
            if val:
                key = str(val).strip().upper()
                if key not in col_map:
                    col_map[key] = idx

        # Group items by (country, ship_date) into shipments
        shipment_groups = defaultdict(list)
        stats = {'processed': 0, 'skipped': 0}

        for row in rows[1:]:
            country = str(row[col_map.get('COUNTRY', 0)] or '').strip().upper()
            m_sku = str(row[col_map.get('M SKU', 2)] or '').strip().upper()
            sku_val = str(row[col_map.get('SKU', 1)] or '').strip()

            if not country or not m_sku or not re.match(r'^M\d+$', m_sku):
                stats['skipped'] += 1
                continue

            stats['processed'] += 1

            shipped = row[9] if len(row) > 9 else 0
            amz_restock = row[col_map.get('AMZ RESTOCK QUANTITY', 6)] if 'AMZ RESTOCK QUANTITY' in col_map else 0
            m_stock = row[col_map.get('M STOCK', 8)] if 'M STOCK' in col_map else 0

            try:
                shipped = int(float(shipped)) if shipped else 0
            except (ValueError, TypeError):
                shipped = 0
            try:
                amz_restock = int(float(amz_restock)) if amz_restock else 0
            except (ValueError, TypeError):
                amz_restock = 0
            try:
                m_stock = int(float(m_stock)) if m_stock else 0
            except (ValueError, TypeError):
                m_stock = 0

            if shipped <= 0:
                stats['skipped'] += 1
                continue

            # Parse date from column 10
            date_text = row[10] if len(row) > 10 else None
            ship_date = parse_ship_date(date_text)

            country_map = {'USA': 'US', 'AUS': 'AU', 'CAN': 'CA'}
            country = country_map.get(country, country)

            shipment_groups[(country, ship_date)].append({
                'm_number': m_sku,
                'sku': sku_val,
                'quantity': shipped,
                'amz_restock': amz_restock,
                'stock_at_ship': m_stock,
            })

        wb.close()

        self.stdout.write(f'Found {len(shipment_groups)} shipment groups from {stats["processed"]} rows')

        if dry_run:
            for (country, date), items in sorted(shipment_groups.items(), key=lambda x: str(x[0][1] or ''))[:10]:
                total = sum(i['quantity'] for i in items)
                self.stdout.write(f'  [DRY] {country} {date}: {len(items)} items, {total} units')
            return

        shipments_created = 0
        items_created = 0

        products_cache = {p.m_number: p for p in Product.objects.all()}

        for (country, ship_date), items in shipment_groups.items():
            shipment = Shipment.objects.create(
                country=country[:10],
                status='shipped',
                shipment_date=ship_date,
                notes=f'Imported from Previous FBA Shipments',
            )
            shipments_created += 1

            for item in items:
                product = products_cache.get(item['m_number'])
                if not product:
                    continue

                ShipmentItem.objects.create(
                    shipment=shipment,
                    product=product,
                    sku=item['sku'][:100],
                    quantity=item['quantity'],
                    quantity_shipped=item['quantity'],
                    amz_restock_quantity=item['amz_restock'],
                    stock_at_ship=item['stock_at_ship'],
                )
                items_created += 1

            shipment.recalculate_totals()

        ImportLog.objects.create(
            import_type='fba_shipments',
            filename=filepath,
            rows_processed=stats['processed'],
            rows_created=items_created,
            rows_updated=0,
            rows_skipped=stats['skipped'],
            errors=[],
        )

        self.stdout.write(self.style.SUCCESS(
            f'FBA Shipments import complete: {shipments_created} shipments, '
            f'{items_created} items created from {stats["processed"]} rows'
        ))
