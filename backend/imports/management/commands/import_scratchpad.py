"""
Import optimal 30-day stock levels from ScratchPad2 sheet.
Clean 2-column lookup: M-number → Optimal Stock Level 30 Days.
"""
import re
from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl import load_workbook
from products.models import Product
from stock.models import StockLevel
from imports.models import ImportLog


class Command(BaseCommand):
    help = 'Import optimal stock levels from ScratchPad2 sheet'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, default=None)
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        filepath = options['file'] or settings.SPREADSHEET_PATH
        dry_run = options['dry_run']

        self.stdout.write(f'Loading {filepath}...')
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb['ScratchPad2']

        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]

        stats = {'processed': 0, 'updated': 0, 'skipped': 0}

        for row in rows[1:]:
            m_number = row[0]
            optimal = row[1]

            if not m_number:
                continue

            m_number = str(m_number).strip().upper()
            if not re.match(r'^M\d+$', m_number):
                continue

            stats['processed'] += 1
            try:
                optimal = int(float(optimal))
            except (ValueError, TypeError):
                optimal = 0

            if dry_run:
                self.stdout.write(f'  [DRY] {m_number}: optimal={optimal}')
                continue

            try:
                product = Product.objects.get(m_number=m_number)
            except Product.DoesNotExist:
                stats['skipped'] += 1
                continue

            stock, _ = StockLevel.objects.get_or_create(product=product)
            stock.optimal_stock_30d = optimal
            stock.stock_deficit = max(0, optimal - stock.current_stock)
            stock.save(update_fields=['optimal_stock_30d', 'stock_deficit', 'updated_at'])
            stats['updated'] += 1

        wb.close()

        if not dry_run:
            ImportLog.objects.create(
                import_type='scratchpad',
                filename=filepath,
                rows_processed=stats['processed'],
                rows_created=0,
                rows_updated=stats['updated'],
                rows_skipped=stats['skipped'],
                errors=[],
            )

        self.stdout.write(self.style.SUCCESS(
            f"ScratchPad2 import complete: {stats['processed']} processed, "
            f"{stats['updated']} updated, {stats['skipped']} skipped"
        ))
