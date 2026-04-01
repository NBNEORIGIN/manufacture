"""
Import FNSKU→SKU→M-number mappings from SKU ASSIGNMENT sheet.
Updates existing SKU records with FNSKU data.
"""
import re
from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl import load_workbook
from products.models import SKU
from imports.models import ImportLog


class Command(BaseCommand):
    help = 'Import FNSKU mappings from SKU ASSIGNMENT sheet'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, default=None)
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        filepath = options['file'] or settings.SPREADSHEET_PATH
        dry_run = options['dry_run']

        self.stdout.write(f'Loading {filepath}...')
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb['SKU ASSIGNMENT']

        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]

        col_map = {}
        for idx, val in enumerate(header):
            if val:
                key = str(val).strip().upper()
                if key not in col_map:
                    col_map[key] = idx

        stats = {'processed': 0, 'updated': 0, 'skipped': 0}

        for row in rows[1:]:
            sku_val = row[col_map.get('SKU', 1)]
            fnsku = row[col_map.get('FNSKU', 2)]

            if not sku_val or not fnsku:
                continue

            sku_val = str(sku_val).strip()
            fnsku = str(fnsku).strip()
            stats['processed'] += 1

            if dry_run:
                self.stdout.write(f'  [DRY] {sku_val} → FNSKU: {fnsku}')
                continue

            updated = SKU.objects.filter(sku=sku_val).update(fnsku=fnsku[:20])
            if updated:
                stats['updated'] += updated
            else:
                stats['skipped'] += 1

        wb.close()

        if not dry_run:
            ImportLog.objects.create(
                import_type='sku_assignment',
                filename=filepath,
                rows_processed=stats['processed'],
                rows_created=0,
                rows_updated=stats['updated'],
                rows_skipped=stats['skipped'],
                errors=[],
            )

        self.stdout.write(self.style.SUCCESS(
            f"SKU ASSIGNMENT import complete: {stats['processed']} processed, "
            f"{stats['updated']} updated, {stats['skipped']} skipped"
        ))
