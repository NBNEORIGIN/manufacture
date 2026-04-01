"""
Import products from MASTER STOCK sheet.
Header at row 2 (0-indexed), rows 0-1 are summary totals.
Key columns: IN PROGRESS, MASTER, DESCRIPTION, BLANK, MATERIAL,
             STOCK, 60 Day Stock, Stock Deficit, Image
"""
import re
from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl import load_workbook
from products.models import Product
from stock.models import StockLevel
from imports.models import ImportLog


class Command(BaseCommand):
    help = 'Import products from MASTER STOCK sheet of the Shipment Stock Sheet'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, default=None)
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        filepath = options['file'] or settings.SPREADSHEET_PATH
        dry_run = options['dry_run']

        self.stdout.write(f'Loading {filepath}...')
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb['MASTER STOCK']

        rows = list(ws.iter_rows(min_row=4, values_only=True))  # skip rows 1-3 (summary + header)
        header_row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]

        col_map = {}
        for idx, val in enumerate(header_row):
            if val:
                col_map[str(val).strip().upper()] = idx

        required = ['MASTER', 'DESCRIPTION', 'BLANK']
        for col in required:
            if col not in col_map:
                self.stderr.write(f'Missing required column: {col}. Found: {list(col_map.keys())}')
                return

        stats = {'processed': 0, 'created': 0, 'updated': 0, 'skipped': 0}
        errors = []

        for row in rows:
            m_number = row[col_map['MASTER']]
            if not m_number or not str(m_number).strip():
                continue

            m_number = str(m_number).strip().upper()
            if not re.match(r'^M\d+$', m_number):
                continue

            stats['processed'] += 1

            description = str(row[col_map['DESCRIPTION']] or '').strip()
            blank = str(row[col_map['BLANK']] or '').strip().upper()
            material = str(row[col_map.get('MATERIAL', -1)] or '').strip() if 'MATERIAL' in col_map else ''
            in_progress = bool(row[col_map.get('IN PROGRESS', -1)]) if 'IN PROGRESS' in col_map else False

            stock_val = row[col_map.get('STOCK', -1)] if 'STOCK' in col_map else 0
            sixty_day = row[col_map.get('60 DAY STOCK', -1)] if '60 DAY STOCK' in col_map else 0
            deficit = row[col_map.get('STOCK DEFICIT', -1)] if 'STOCK DEFICIT' in col_map else 0

            def safe_int(v):
                try:
                    return int(float(v))
                except (ValueError, TypeError):
                    return 0

            stock_val = safe_int(stock_val)
            sixty_day = safe_int(sixty_day)
            deficit = safe_int(deficit)

            if not description:
                stats['skipped'] += 1
                continue

            if dry_run:
                self.stdout.write(f'  [DRY] {m_number}: {description[:50]} ({blank})')
                continue

            product, created = Product.objects.update_or_create(
                m_number=m_number,
                defaults={
                    'description': description[:500],
                    'blank': blank[:50],
                    'material': material[:100],
                    'in_progress': in_progress,
                    'active': True,
                },
            )

            StockLevel.objects.update_or_create(
                product=product,
                defaults={
                    'current_stock': stock_val,
                    'sixty_day_sales': sixty_day,
                    'stock_deficit': max(0, deficit),
                },
            )

            if created:
                stats['created'] += 1
            else:
                stats['updated'] += 1

        wb.close()

        if not dry_run:
            ImportLog.objects.create(
                import_type='master_stock',
                filename=filepath,
                rows_processed=stats['processed'],
                rows_created=stats['created'],
                rows_updated=stats['updated'],
                rows_skipped=stats['skipped'],
                errors=errors,
            )

        self.stdout.write(self.style.SUCCESS(
            f"MASTER STOCK import complete: {stats['processed']} processed, "
            f"{stats['created']} created, {stats['updated']} updated, "
            f"{stats['skipped']} skipped"
        ))
