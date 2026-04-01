"""
Import SKU→M-number mappings from ASSEMBLY sheet.
Maps marketplace SKUs to master M-numbers with channel, ASIN, blank, personalisation flag.
"""
import re
from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl import load_workbook
from products.models import Product, SKU
from imports.models import ImportLog


CHANNEL_MAP = {
    'UK': 'UK', 'EBAY': 'EBAY', 'ETSY': 'ETSY', 'AMAZON': 'UK',
    'FR': 'FR', 'DE': 'DE', 'USA': 'US', 'US': 'US',
    'CA': 'CA', 'CAN': 'CA', 'AU': 'AU', 'AUS': 'AU',
    'ES': 'ES', 'IT': 'IT', 'NL': 'NL', 'SHOPIFY': 'SHOPIFY',
    'FR CRAFTS': 'FR_CRAFTS', 'FR DESIGNED': 'FR_DESIGNED',
    'IT DESIGNED': 'IT_DESIGNED',
}

SKIP_COUNTRIES = {
    'OLD LISTING', 'REUSE', 'STOCK',
    'FR CRAFTS - DUPLICATE DESIGNED', 'FR DESIGNED - DUPLICATE CRAFTS',
}


class Command(BaseCommand):
    help = 'Import SKU mappings from ASSEMBLY sheet'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, default=None)
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        filepath = options['file'] or settings.SPREADSHEET_PATH
        dry_run = options['dry_run']

        self.stdout.write(f'Loading {filepath}...')
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb['ASSEMBLY']

        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]

        col_map = {}
        for idx, val in enumerate(header):
            if val:
                key = str(val).strip().upper()
                if key not in col_map:  # first occurrence only — right side has lookup dupes
                    col_map[key] = idx

        stats = {'processed': 0, 'created': 0, 'updated': 0, 'skipped': 0}
        errors = []
        unknown_products = set()

        for row in rows[1:]:
            sku_val = row[col_map.get('SKU', 0)]
            master_sku = row[col_map.get('MASTER SKU', 1)]

            if not sku_val or not master_sku:
                continue
            sku_val = str(sku_val).strip()
            master_sku = str(master_sku).strip().upper()

            if not re.match(r'^M\d+$', master_sku):
                continue

            stats['processed'] += 1

            country_raw = str(row[col_map.get('COUNTRY', 3)] or '').strip().upper()
            if country_raw in SKIP_COUNTRIES or not country_raw:
                stats['skipped'] += 1
                continue
            channel = CHANNEL_MAP.get(country_raw, country_raw)
            # Skip entries with garbage country values (e.g. contains M-number references)
            if len(channel) > 20:
                stats['skipped'] += 1
                continue

            asin = str(row[col_map.get('ASIN', -1)] or '').strip() if 'ASIN' in col_map else ''
            new_sku = str(row[col_map.get('NEW SKU', -1)] or '').strip() if 'NEW SKU' in col_map else ''
            is_personalised = str(row[col_map.get('IS PERSONALISED?', -1)] or '').strip().upper()
            is_personalised = is_personalised in ('YES', 'TRUE', '1', 'Y')

            try:
                product = Product.objects.get(m_number=master_sku)
            except Product.DoesNotExist:
                if master_sku not in unknown_products:
                    unknown_products.add(master_sku)
                    errors.append(f'Unknown product: {master_sku}')
                stats['skipped'] += 1
                continue

            if is_personalised and not product.is_personalised:
                product.is_personalised = True
                product.save(update_fields=['is_personalised'])

            if dry_run:
                self.stdout.write(f'  [DRY] {sku_val} ({channel}) → {master_sku}')
                continue

            _, created = SKU.objects.update_or_create(
                sku=sku_val,
                channel=channel,
                defaults={
                    'product': product,
                    'new_sku': new_sku[:100],
                    'asin': asin[:20],
                    'active': True,
                },
            )

            if created:
                stats['created'] += 1
            else:
                stats['updated'] += 1

        wb.close()

        if unknown_products:
            self.stdout.write(self.style.WARNING(f'{len(unknown_products)} unknown M-numbers (run import_master_stock first)'))

        if not dry_run:
            ImportLog.objects.create(
                import_type='assembly',
                filename=filepath,
                rows_processed=stats['processed'],
                rows_created=stats['created'],
                rows_updated=stats['updated'],
                rows_skipped=stats['skipped'],
                errors=errors,
            )

        self.stdout.write(self.style.SUCCESS(
            f"ASSEMBLY import complete: {stats['processed']} processed, "
            f"{stats['created']} created, {stats['updated']} updated, "
            f"{stats['skipped']} skipped"
        ))
