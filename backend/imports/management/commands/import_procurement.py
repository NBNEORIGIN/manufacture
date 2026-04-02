"""
Import materials from PROCUREMENT sheet.
21 rows, well-structured: MaterialID, MaterialName, Category, etc.
"""
from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl import load_workbook
from procurement.models import Material
from imports.models import ImportLog


def safe_str(row, idx, max_len=200):
    if idx >= len(row) or row[idx] is None:
        return ''
    return str(row[idx]).strip()[:max_len]


def safe_int(row, idx):
    if idx >= len(row) or row[idx] is None:
        return 0
    try:
        return int(float(row[idx]))
    except (ValueError, TypeError):
        return 0


class Command(BaseCommand):
    help = 'Import materials from PROCUREMENT sheet'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, default=None)
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        filepath = options['file'] or settings.SPREADSHEET_PATH
        dry_run = options['dry_run']

        self.stdout.write(f'Loading {filepath}...')
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb['PROCUREMENT']

        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]

        col = {}
        for idx, val in enumerate(header):
            if val:
                col[str(val).strip()] = idx

        stats = {'processed': 0, 'created': 0, 'updated': 0}

        for row in rows[1:]:
            if col.get('MaterialID', 0) >= len(row):
                continue
            mat_id = row[col.get('MaterialID', 0)]
            if not mat_id:
                continue

            mat_id = str(mat_id).strip()
            stats['processed'] += 1

            name = safe_str(row, col.get('MaterialName', 1))
            if not name:
                continue

            price = None
            price_idx = col.get('CurrentPrice', 13)
            if price_idx < len(row) and row[price_idx] is not None:
                try:
                    price = float(row[price_idx])
                except (ValueError, TypeError):
                    pass

            if dry_run:
                self.stdout.write(f'  [DRY] {mat_id}: {name}')
                continue

            _, created = Material.objects.update_or_create(
                material_id=mat_id,
                defaults={
                    'name': name,
                    'category': safe_str(row, col.get('Category', 2), 100),
                    'unit_of_measure': safe_str(row, col.get('UnitOfMeasure', 3), 50),
                    'current_stock': safe_int(row, col.get('CurrentStock', 4)),
                    'reorder_point': safe_int(row, col.get('ReorderPoint', 5)),
                    'standard_order_quantity': safe_int(row, col.get('StandardOrderQuantity', 6)),
                    'preferred_supplier': safe_str(row, col.get('PreferredSupplierID', 7)),
                    'product_page_url': safe_str(row, col.get('ProductPageURL', 8), 500),
                    'lead_time_days': safe_int(row, col.get('LeadTimeDays', 9)),
                    'safety_stock': safe_int(row, col.get('SafetyStockQuantity', 10)),
                    'in_house_description': safe_str(row, col.get('In-House description', 11)),
                    'notes': safe_str(row, col.get('Notes', 12), 1000),
                    'current_price': price,
                },
            )

            if created:
                stats['created'] += 1
            else:
                stats['updated'] += 1

        wb.close()

        if not dry_run:
            ImportLog.objects.create(
                import_type='procurement',
                filename=filepath,
                rows_processed=stats['processed'],
                rows_created=stats['created'],
                rows_updated=stats['updated'],
                rows_skipped=0,
                errors=[],
            )

        self.stdout.write(self.style.SUCCESS(
            f"PROCUREMENT import complete: {stats['processed']} processed, "
            f"{stats['created']} created, {stats['updated']} updated"
        ))
