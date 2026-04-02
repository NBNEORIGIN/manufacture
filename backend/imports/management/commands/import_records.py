"""
Import production records from RECORDS sheet.
Columns: Date, WEEK, SKU, M NUMBER, NUMBER PRINTED, ERRORS,
         TOTAL MADE, PROCESS, FAILURE REASON, CORRECTION?
"""
import re
from datetime import datetime
from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl import load_workbook
from products.models import Product, SKU
from production.models_records import ProductionRecord
from imports.models import ImportLog


class Command(BaseCommand):
    help = 'Import production records from RECORDS sheet'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, default=None)
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        filepath = options['file'] or settings.SPREADSHEET_PATH
        dry_run = options['dry_run']

        self.stdout.write(f'Loading {filepath}...')
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb['RECORDS']

        rows = list(ws.iter_rows(values_only=True))
        stats = {'processed': 0, 'created': 0, 'skipped': 0}

        # Build product lookup caches
        products_by_m = {p.m_number: p for p in Product.objects.all()}
        sku_to_product = {}
        for s in SKU.objects.select_related('product').all():
            sku_to_product[s.sku] = s.product

        for row in rows[1:]:
            if len(row) < 5:
                continue

            date_val = row[0]
            if not date_val or not isinstance(date_val, datetime):
                stats['skipped'] += 1
                continue

            date = date_val.date() if isinstance(date_val, datetime) else date_val
            week = row[1] if len(row) > 1 else None
            sku_val = str(row[2] or '').strip() if len(row) > 2 else ''
            m_number = str(row[3] or '').strip().upper() if len(row) > 3 else ''
            printed = row[4] if len(row) > 4 else 0
            errors = row[5] if len(row) > 5 else 0
            total = row[6] if len(row) > 6 else 0
            machine = str(row[7] or '').strip().upper() if len(row) > 7 else ''
            reason = str(row[8] or '').strip() if len(row) > 8 else ''
            correction = str(row[9] or '').strip() if len(row) > 9 else ''

            if not sku_val and not m_number:
                stats['skipped'] += 1
                continue

            stats['processed'] += 1

            def safe_int(v):
                try:
                    return int(float(v)) if v else 0
                except (ValueError, TypeError):
                    return 0

            printed = safe_int(printed)
            errors = safe_int(errors)
            total = safe_int(total) or printed

            # Resolve product
            product = None
            if m_number and re.match(r'^M\d+$', m_number):
                product = products_by_m.get(m_number)
            if not product and sku_val:
                product = sku_to_product.get(sku_val)

            try:
                week_num = int(float(week)) if week else None
            except (ValueError, TypeError):
                week_num = None

            if dry_run:
                err_flag = f' ({errors} errors: {reason})' if errors else ''
                self.stdout.write(f'  [DRY] {date} {sku_val or m_number} {machine} {printed} printed{err_flag}')
                if stats['processed'] >= 10:
                    self.stdout.write('  ... (showing first 10)')
                    break
                continue

            ProductionRecord.objects.create(
                date=date,
                week_number=week_num,
                product=product,
                sku=sku_val[:100],
                number_printed=printed,
                errors=errors,
                total_made=total,
                machine=machine[:50],
                failure_reason=reason,
                correction=correction,
            )
            stats['created'] += 1

        wb.close()

        if not dry_run:
            ImportLog.objects.create(
                import_type='records',
                filename=filepath,
                rows_processed=stats['processed'],
                rows_created=stats['created'],
                rows_updated=0,
                rows_skipped=stats['skipped'],
                errors=[],
            )

        self.stdout.write(self.style.SUCCESS(
            f"RECORDS import complete: {stats['processed']} processed, "
            f"{stats['created']} created, {stats['skipped']} skipped"
        ))
